from collections import defaultdict
from fastapi import FastAPI
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
import shutil
import os
import pythoncom
import win32com.client as win32
import requests
import traceback

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

PLANTILLA = "Manifiesto_Preview.xlsx"

CAPACIDADES = {
    "tote": "1000 litros",
    "tambo": "200 litros",
    "paca": "600 kg",
    "pieza": "1",
    "tarima": "1"
}

@app.get("/generar-manifiesto")
def generar_manifiesto():
    if not os.path.exists(PLANTILLA):
        return JSONResponse(status_code=404, content={"detail": "Plantilla no encontrada."})

    try:
        response = requests.get("http://localhost:4002/api/manifiesto-temporal")
        if response.status_code != 200:
            raise Exception("No se pudieron obtener las filas temporales")
        datos_json = response.json()

        agrupados = defaultdict(lambda: {
            "codigo": "",
            "contenedor": "",
            "cantidad": 0.0,
            "peso": 0.0
        })

        for fila in datos_json:
            nombre = fila["nombre"]
            contenedor = fila.get("contenedor", "").lower()
            agrupados[nombre]["codigo"] = fila.get("codigo", "")
            agrupados[nombre]["contenedor"] = contenedor
            agrupados[nombre]["cantidad"] += float(fila.get("cantidad", 0))
            agrupados[nombre]["peso"] += float(fila.get("peso", 0))

        datos = [
            [nombre, info["codigo"], info["contenedor"],
             info["cantidad"], info["peso"],
             CAPACIDADES.get(info["contenedor"], "N/A")]
            for nombre, info in agrupados.items()
        ]

        if not datos:
            raise Exception("No hay datos para generar el manifiesto")

        gen_resp = requests.post("http://localhost:4002/api/documento-final")
        if gen_resp.status_code != 201:
            raise Exception("No se pudo registrar el manifiesto en la base de datos")

        nombre_archivo = gen_resp.json()["archivo"]
        num_manifiesto = nombre_archivo.replace(".xlsx", "")
        excel_filename = f"Manifiesto {num_manifiesto} SAI.xlsx"

        shutil.copyfile(PLANTILLA, excel_filename)
        wb = load_workbook(excel_filename)

        for idx in range(4):
            ws = wb.worksheets[idx]
            ws["U12"] = num_manifiesto
            fila_inicio = 20
            for i, fila in enumerate(datos):
                f = fila_inicio + i
                ws[f"B{f}"] = fila[0]  # nombre
                ws[f"P{f}"] = fila[5]  # capacidad
                ws[f"S{f}"] = fila[2]  # contenedor

                peso = float(str(fila[3]).replace(",", "."))
                cantidad = float(str(fila[4]).replace(",", "."))

                ws[f"V{f}"] = peso
                ws[f"V{f}"].number_format = '0'
                ws[f"Y{f}"] = cantidad
                ws[f"Y{f}"].number_format = '0.00'

        wb.save(excel_filename)
        wb.close()

        return JSONResponse(
            content={"detail": "Manifiesto generado correctamente", "archivo": nombre_archivo},
            status_code=200
        )

    except Exception as e:
        print("❌ Error en /generar-manifiesto:\n", traceback.format_exc())
        return JSONResponse(status_code=500, content={"detail": f"Error interno: {str(e)}"})


@app.get("/preview-manifiesto")
def preview_manifiesto():
    try:
        print("✅ Iniciando generación del PDF")

        resp = requests.get("http://localhost:4002/api/documentos/ultimo")
        if resp.status_code != 200:
            return JSONResponse(status_code=404, content={"detail": "No se encontró el número del manifiesto."})

        nombre_archivo = resp.json()["nombre_archivo"].replace(".xlsx", "")
        excel_filename = f"Manifiesto {nombre_archivo} SAI.xlsx"
        pdf_filename = f"Manifiesto {nombre_archivo} SAI.pdf"

        if not os.path.exists(excel_filename):
            return JSONResponse(status_code=404, content={"detail": "Excel no encontrado."})

        pythoncom.CoInitialize()
        import win32com.client.gencache
        win32com.client.gencache.is_readonly = False
        win32com.client.gencache.Rebuild()
        excel = win32.gencache.EnsureDispatch("Excel.Application")
        wb = excel.Workbooks.Open(os.path.abspath(excel_filename))
        excel.CalculateFullRebuild()
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_filename))
        wb.Close(False)
        excel.Quit()
        pythoncom.CoUninitialize()

        return FileResponse(
            path=pdf_filename,
            filename=pdf_filename,
            media_type="application/pdf"
        )

    except Exception as e:
        print(f"❌ Error al generar PDF: {e}")
        return JSONResponse(status_code=500, content={"detail": f"Error al generar PDF: {str(e)}"})


@app.delete("/terminar-manifiesto")
def terminar_manifiesto():
    try:
        borrar_resp = requests.delete("http://localhost:4002/api/manifiesto-temporal")
        if borrar_resp.status_code != 200:
            return JSONResponse(status_code=500, content={"detail": "Error al limpiar manifiesto temporal."})
        return JSONResponse(content={"detail": "Manifiesto terminado y temporal limpiado."})
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": f"Error: {str(e)}"})


@app.get("/descargar-excel/{nombre_archivo}")
def descargar_excel(nombre_archivo: str):
    excel_filename = f"Manifiesto {nombre_archivo} SAI.xlsx"
    if not os.path.exists(excel_filename):
        return JSONResponse(status_code=404, content={"detail": "Excel no encontrado."})
    return FileResponse(excel_filename, filename=excel_filename)


@app.get("/descargar-pdf/{nombre_archivo}")
def descargar_pdf(nombre_archivo: str):
    pdf_filename = f"Manifiesto {nombre_archivo} SAI.pdf"
    if not os.path.exists(pdf_filename):
        return JSONResponse(status_code=404, content={"detail": "PDF no encontrado."})
    return FileResponse(pdf_filename, filename=pdf_filename)
